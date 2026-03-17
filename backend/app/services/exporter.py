from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.batch import Batch
from app.models.export_job import ExportJob
from app.models.extracted_row import ExtractedRow


SUMMARY_COLUMNS = [
    "source_file",
    "date",
    "agent",
    "customer",
    "sourcing_executive",
    "weave",
    "quality",
    "shafts",
    "marketing_executive",
    "buyer_reference_no",
    "design_no",
    "is_warp_butta",
    "is_weft_butta",
    "is_warp2_sizing_count",
    "is_seersucker",
    "warp_count",
    "warp_rate_per_kg",
    "warp_rate_incl_gst",
    "warp_gst",
    "warp_content",
    "warp_yarn_type",
    "warp_mill",
    "warp_epi_on_loom",
    "weft_count",
    "weft_rate_per_kg",
    "weft_rate_incl_gst",
    "weft_gst",
    "weft_content",
    "weft_yarn_type",
    "weft_mill",
    "weft_ppi",
    "grey_width",
    "epi_on_table",
    "meters_per_120_yards",
    "total_ends",
    "epi_difference",
    "reed_space",
    "warp_crimp_percent",
    "weight_warp1",
    "cost_warp1",
    "composition_warp1",
    "weight_weft1",
    "cost_weft1",
    "composition_weft1",
    "gsm_total_yarn_cost",
    "fabric_total_yarn_cost",
    "fabric_weight_glm_inc_sizing",
    "sizing_per_kg_rate",
    "sizing_per_kg_cost",
    "weaving_charges_rate",
    "weaving_charges_cost",
    "freight_rate",
    "freight_cost",
    "butta_cutting_rate",
    "butta_cutting_cost",
    "yarn_wastage_rate",
    "yarn_wastage_cost",
    "value_loss_interest_rate",
    "value_loss_interest_cost",
    "payment_term",
    "particulars_total_cost",
    "commission_cd_rate",
    "commission_cd_cost",
    "remark",
    "other_cost_if_any_rate",
    "other_cost_if_any_remarks",
    "extra_remarks_if_any",
    "total_price",
    "target_price",
    "weaving_charge_as_per_tp",
    "order_quantity",
    "yarn_requirement_warp1",
    "yarn_requirement_weft1",
    "yarn_requirement_total",
    "cover_factor",
    "batch_id",
    "reviewed_by",
    "reviewed_at",
]


def create_export(db: Session, batch: Batch, created_by_id: int) -> ExportJob:
    export = ExportJob(batch_id=batch.id, created_by_id=created_by_id, status="processing")
    db.add(export)
    db.commit()
    db.refresh(export)

    file_path = settings.export_root / f"batch_{batch.id}_export_{export.id}.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _build_workbook(db, batch, file_path)

    export.file_path = str(file_path.resolve())
    export.status = "completed"
    batch.export_status = "ready"
    db.commit()
    db.refresh(export)
    return export


def _build_workbook(db: Session, batch: Batch, file_path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(SUMMARY_COLUMNS)

    rows = db.query(ExtractedRow).filter(ExtractedRow.batch_id == batch.id).all()
    for row in rows:
        payload = row.reviewed_data or row.data
        values = [payload.get(column) for column in SUMMARY_COLUMNS[1:-3]]
        summary_sheet.append(
            [
                row.source_file.original_name,
                *values,
                batch.id,
                row.last_edited_by.name if row.last_edited_by else None,
                row.last_edited_at.isoformat() if row.last_edited_at else None,
            ]
        )

    audit_sheet = workbook.create_sheet("Audit")
    audit_sheet.append(
        [
            "row_id",
            "source_file",
            "status",
            "field_name",
            "raw_text",
            "normalized_value",
            "confidence",
            "validation_issues",
        ]
    )
    for row in rows:
        for audit in row.audits:
            audit_sheet.append(
                [
                    row.id,
                    row.source_file.original_name,
                    row.status,
                    audit.field_name,
                    audit.raw_text,
                    audit.normalized_value,
                    audit.confidence,
                    ", ".join(audit.validation_issues),
                ]
            )

    error_sheet = workbook.create_sheet("Errors")
    error_sheet.append(["source_file", "status", "error_message"])
    for uploaded_file in batch.files:
        if uploaded_file.status == "failed":
            error_sheet.append([uploaded_file.original_name, uploaded_file.status, uploaded_file.error_message])

    workbook.save(file_path)
